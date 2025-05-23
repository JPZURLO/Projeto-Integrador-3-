from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import jwt
import mysql.connector
from datetime import datetime, timedelta
import os
from datetime import datetime, timedelta, timezone
import traceback
import json  # Adicione esta linha
from openpyxl import load_workbook
from functools import wraps  # Import the wraps decorator
from werkzeug.utils import secure_filename  # Adicione esta linha
app = Flask(__name__, static_folder='FRONT/html', static_url_path='/FRONT/html')
CORS(app) 
secretKey = 'suaChaveSecretaAqui'  # Troque por uma chave mais forte

# Configuração do diretório de upload
app.config['UPLOAD_FOLDER'] = 'uploads' # Adicione esta linha

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ['png', 'jpg', 'jpeg', 'gif']

# Configuração do banco de dados
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': os.getenv('DB_PASSWORD', 'De182246@'),  # Melhor usar variável de ambiente
    'database': 'gestaopublicadigital'
}

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token or not token.startswith('Bearer '):
            return jsonify({'message': 'Token não fornecido!'}), 401
        try:
            token = token.split(' ')[1]
            data = jwt.decode(token, secretKey, algorithms=['HS256'])
            # Você pode adicionar informações do usuário ao contexto se necessário
            return f(*args, **kwargs)
        except:
            return jsonify({'message': 'Token inválido!'}), 401
    return decorated

@app.route('/api/notifications', methods=['GET'])
@token_required
def get_notifications():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        now = datetime.now().date()
        five_days_later = now + timedelta(days=5)

        notifications = []

        # Notificações de obras com data final próxima (5 dias)
        cursor.execute("""
            SELECT NomeDaObra, DataDeEntrega
            FROM obras
            WHERE DataDeEntrega <= %s AND DataDeEntrega >= %s
        """, (five_days_later, now))
        obras_terminando = cursor.fetchall()
        for obra in obras_terminando:
            notifications.append({
                'type': 'deadline',
                'message': f'A obra "{obra["NomeDaObra"]}" está terminando em {obra["DataDeEntrega"].strftime("%d/%m/%Y")}.'
            })

        # Notificações de novas obras cadastradas (exemplo: últimas 5 obras)
        cursor.execute("""
            SELECT NomeDaObra, DataDeInicio
            FROM obras
            ORDER BY id DESC
            LIMIT 5
        """)
        novas_obras = cursor.fetchall()
        for obra in novas_obras:
            notifications.append({
                'type': 'new_obra',
                'message': f'Nova obra cadastrada: "{obra["NomeDaObra"]}" em {obra["DataDeInicio"].strftime("%d/%m/%Y")}.'
            })

        cursor.close()
        conn.close()

        return jsonify({'notifications': notifications})

    except Exception as e:
        print(f"Erro ao buscar notificações: {e}")
        return jsonify({'error': 'Erro ao buscar notificações'}), 500

@app.route('/api/obterNomeCompleto', methods=['GET'])
def obter_nome_completo():
    token = request.headers.get('Authorization')
    if not token or not token.startswith('Bearer '):
        return jsonify({'error': 'Token não fornecido ou inválido'}), 401

    token = token.split(' ')[1]
    try:
        payload = jwt.decode(token, secretKey, algorithms=['HS256'])
        user_id = payload.get('id')
        if not user_id:
            return jsonify({'error': 'ID de usuário não encontrado no token'}), 401

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT NomeCompleto FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if user:
            return jsonify({'NomeCompleto': user['NomeCompleto']})
        else:
            return jsonify({'error': 'Usuário não encontrado'}), 404

    except jwt.ExpiredSignatureError:
        return jsonify({'error': 'Token expirado'}), 401
    except jwt.InvalidTokenError:
        return jsonify({'error': 'Token inválido'}), 401
    except Exception as e:
        print(f"Erro ao obter nome completo: {e}")
        return jsonify({'error': 'Erro interno do servidor'}), 500

@app.route('/obras-dados')
def get_obras_dados():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("SELECT COUNT(*) as total FROM obras")
        total_obras = cursor.fetchone()['total']

        cursor.execute("""
            SELECT c.TipoDeObra, COUNT(*) as quantidade
            FROM obras o
            JOIN classificacaodasobras c ON o.ClassificacaoDaObra = c.Id
            GROUP BY c.TipoDeObra
        """)
        obras_por_tipo = cursor.fetchall()

        cursor.execute("""
            SELECT r.Nome_Regiao, COUNT(*) as quantidade
            FROM obras o
            JOIN regioesbrasil r ON o.Regiao = r.Id
            GROUP BY r.Nome_Regiao
        """)
        obras_por_regiao = cursor.fetchall()

        cursor.execute("""
            SELECT r.Nome_Regiao, c.TipoDeObra, COUNT(*) as quantidade
            FROM obras o
            JOIN regioesbrasil r ON o.Regiao = r.Id
            JOIN classificacaodasobras c ON o.ClassificacaoDaObra = c.Id
            GROUP BY r.Nome_Regiao, c.TipoDeObra
        """)
        total_por_regiao_tipo = cursor.fetchall()

        cursor.close()
        conn.close()

        return jsonify({
            'totalObras': total_obras,
            'obrasPorTipo': obras_por_tipo,
            'obrasPorRegiao': obras_por_regiao,
            'totalPorRegiaoTipo': total_por_regiao_tipo
        })
    except Exception as e:
        print(f"Erro ao buscar dados: {e}")
        return jsonify({'error': 'Erro ao buscar dados'}), 500


def get_db_connection():
    return mysql.connector.connect(**db_config)

import hashlib

def gerar_hash_imagem(imagens):
    hashes = []
    for imagem_content in imagens:
        hash_md5 = hashlib.md5(imagem_content).hexdigest()
        hashes.append(hash_md5)
    return hashes

def get_db_connection():
    return mysql.connector.connect(**db_config)

#  Rota de login
@app.route('/login', methods=['POST'])
def login():
    print(">>> Rota /login acessada")
    try:
        print(">>> Tentando obter dados JSON")
        data = request.get_json()
        print(f">>> Dados recebidos: {data}")
        email = data.get('Email')
        senha = data.get('Senha')
        print(f">>> Email: {email}, Senha: {senha}")

        print(">>> Tentando conectar ao banco de dados")
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        print(">>> Conexão ao banco de dados bem-sucedida")

        print(">>> Executando a consulta SQL")
        cursor.execute('''
            SELECT u.*, t.tipo AS TipoDeUsuario
            FROM users u
            JOIN tipousuario t ON u.TipoUsuarioId = t.id
            WHERE u.Email = %s AND u.Senha = %s
        ''', (email, senha))
        user = cursor.fetchone()
        print(f">>> Resultado da consulta: {user}")

        cursor.close()
        db.close()
        print(">>> Conexão ao banco de dados fechada")

        if user:
            print(">>> Usuário encontrado, criando payload JWT")
            payload = {
                'id': user['id'],
                'exp': datetime.now(timezone.utc) + timedelta(hours=1),
                'user_type': user['TipoDeUsuario']
            }
            token = jwt.encode(payload, secretKey, algorithm='HS256')
            print(f">>> Token JWT criado: {token}")
            print(">>> Retornando resposta de sucesso")
            return jsonify({
                'success': True,
                'message': 'Login realizado com sucesso!',
                'token': token,
                'NomeCompleto': user['NomeCompleto'],
                'TipoDeUsuario': user['TipoDeUsuario']
            })
        else:
            print(">>> Usuário não encontrado, retornando erro de credenciais")
            return jsonify({'success': False, 'message': 'Credenciais inválidas.'}), 401
    except Exception as e:
        print(f">>> OCORREU UM ERRO: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500 # Garanta que o erro seja retornado como JSON

import hashlib

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token or not token.startswith('Bearer '):
            return jsonify({'message': 'Token não fornecido!'}), 401
        try:
            token = token.split(' ')[1]
            data = jwt.decode(token, secretKey, algorithms=['HS256'])
            current_user_type = data['user_type']
        except:
            return jsonify({'message': 'Token inválido!'}), 401
        return f(current_user_type, *args, **kwargs)
    return decorated

def admin_gestor_engenheiro_required(f):
    @wraps(f)
    def decorated(user_type, *args, **kwargs):
        if user_type.lower() not in ['administrador', 'gestor', 'engenheiro']:
            return jsonify({'message': 'Permissão negada!'}), 403
        return f(user_type, *args, **kwargs)
    return decorated 

#  Rota para obter opções de cadastro da obra
# Rota para obter opções de cadastro da obra
@app.route('/cadastrar-obra', methods=['GET'])
def cadastrar_obra():
    """Rota para buscar os dados de regiões, classificações, status e estados."""
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        # Buscar regiões SEM duplicatas
        cursor.execute('SELECT DISTINCT id, Nome_Regiao FROM regioesbrasil')
        regioes = cursor.fetchall()

        # Buscar classificações de obra
        cursor.execute('SELECT DISTINCT id, TipoDeObra FROM classificacaodasobras')
        classificacoes = cursor.fetchall()

        # Buscar status da obra
        cursor.execute('SELECT DISTINCT id, Classificacao FROM statusdaobra')
        status = cursor.fetchall()
        
        # Buscar estados
        cursor.execute('SELECT id, nome FROM estados')
        estados = cursor.fetchall()
        
        # Buscar cidades de todos os estados
        cursor.execute('SELECT id, nome, estado_id FROM cidades')
        cidades = cursor.fetchall()

        cursor.close()
        db.close()

        # Organizar cidades por estado
        cidades_por_estado = {}
        for cidade in cidades:
            estado_id = cidade['estado_id']
            if estado_id not in cidades_por_estado:
                cidades_por_estado[estado_id] = []
            cidades_por_estado[estado_id].append({'id': cidade['id'], 'nome': cidade['nome']})

        return jsonify({
            'regioes': regioes,
            'classificacoes': classificacoes,
            'status': status,
            'estados': estados,
            'cidades': cidades_por_estado # Retorna um dicionário de cidades por estado
        })
    except Exception as e:
        print(f"Erro em /cadastrar-obra: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/obter-cidades/<int:estado_id>', methods=['GET'])
@token_required
def obter_cidades(current_user_type, estado_id):
    """Rota para buscar as cidades de um estado específico."""
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)
        cursor.execute('SELECT id, nome FROM cidades WHERE estado_id = %s ORDER BY nome', (estado_id,))
        cidades = cursor.fetchall()
        cursor.close()
        db.close()
        return jsonify({'cidades': cidades})
    except Exception as e:
        print(f"Erro em /obter-cidades/{estado_id}: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/adicionar-obra', methods=['POST'])
def adicionar_obra():
    """Rota para adicionar uma nova obra ao banco de dados."""
    print("Início da função adicionar_obra")
    try:
        print("Dentro do bloco try")
        print("Dados recebidos do formulário:")
        print(request.form)

        nome_da_obra = request.form.get('NomeDaObra')
        regiao_nome = request.form.get('Regiao')
        estado_id = request.form.get('Estado')  # Recebe o ID do estado
        cidade_id = request.form.get('Cidade')  # Recebe o ID da cidade
        classificacao_nome = request.form.get('ClassificacaoDaObra')
        status_nome = request.form.get('Status')
        data_inicio = request.form.get('DataDeInicio')
        data_termino = request.form.get('DataDeEntrega')
        orcamento_utilizado = request.form.get('orcamento')
        descricao = request.form.get('Descricao')
        engenheiro_responsavel = request.form.get('EngResponsavel')

        # Validação dos dados recebidos
        if not all([nome_da_obra, regiao_nome, estado_id, cidade_id, classificacao_nome, status_nome, data_inicio, data_termino, orcamento_utilizado, descricao, engenheiro_responsavel]):
            return jsonify({'error': 'Todos os campos são obrigatórios!'}), 400

        # Validação do orçamento
        orcamento_utilizado = orcamento_utilizado.replace('.', '').replace(',', '.')
        try:
            orcamento_utilizado = float(orcamento_utilizado)
        except ValueError:
            return jsonify({'error': 'O valor do orçamento é inválido!'}), 400

        # Validação das datas
        try:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            data_termino_obj = datetime.strptime(data_termino, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de data inválido. Use o formato AAAA-MM-DD'}), 400

        if data_termino_obj < data_inicio_obj:
            return jsonify({'error': 'A data de término não pode ser anterior à data de início.'}), 400

        print("Valores dos campos:")
        print(f"Nome da obra: {nome_da_obra}")
        print(f"Região: {regiao_nome}")
        print(f"Estado: {estado_id}")
        print(f"Cidade: {cidade_id}")
        print(f"Classificação: {classificacao_nome}")
        print(f"Status: {status_nome}")
        print(f"Data de início: {data_inicio}")
        print(f"Data de término: {data_termino}")
        print(f"Orcamento utilizado: {orcamento_utilizado}")
        print(f"Descrição: {descricao}")
        print(f"Engenheiro responsável: {engenheiro_responsavel}")

        imagens = request.files.getlist('imagem')
        imagem_paths = []
        imagens_hash_existente = set()  # Conjunto para verificar imagens duplicadas no request

        for imagem in imagens:
            if imagem and allowed_file(imagem.filename):
                # Calcula o hash da imagem para verificar duplicatas
                imagem_hash = hashlib.md5(imagem.read()).hexdigest()
                if imagem_hash in imagens_hash_existente:
                    return jsonify({'error': 'Uma ou mais imagens enviadas são duplicadas nesta requisição.'}), 400
                imagens_hash_existente.add(imagem_hash)  # Adiciona ao conjunto de hashes da requisição
                imagem.seek(0)  # Reseta o ponteiro do arquivo para salvar a imagem
                filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{imagem.filename}"
                imagem_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                imagem.save(imagem_path)
                imagem_paths.append(f"/uploads/{filename}")
            elif imagem:
                return jsonify({'error': 'Tipo de arquivo não permitido para a imagem'}), 400

        db = None
        cursor = None
        try:
            db = get_db_connection()
            cursor = db.cursor()

            # Buscar IDs de FKs (estado, cidade, regiao, classificacao, status)
            cursor.execute("SELECT id FROM regioesbrasil WHERE id = %s", (regiao_nome,))
            regiao_id = cursor.fetchone()
            if not regiao_id:
                return jsonify({'error': 'Região não encontrada'}), 400
            regiao_id = regiao_id[0]

            cursor.execute("SELECT id FROM classificacaodasobras WHERE id = %s", (classificacao_nome,))
            classificacao_id = cursor.fetchone()
            if not classificacao_id:
                return jsonify({'error': 'Classificação da obra não encontrada'}), 400
            classificacao_id = classificacao_id[0]

            cursor.execute("SELECT id FROM statusdaobra WHERE id = %s", (status_nome,))
            status_id = cursor.fetchone()
            if not status_id:
                return jsonify({'error': 'Status não encontrado'}), 400
            status_id = status_id[0]

            # Verificar se o estado e cidade existem
            cursor.execute("SELECT id FROM estados WHERE id = %s", (estado_id,))
            estado_existe = cursor.fetchone()
            if not estado_existe:
                return jsonify({'error': 'Estado não encontrado'}), 400

            cursor.execute("SELECT id FROM cidades WHERE id = %s", (cidade_id,))
            cidade_existe = cursor.fetchone()
            if not cidade_existe:
                return jsonify({'error': 'Cidade não encontrada'}), 400

            import json
            cursor.execute("""
                INSERT INTO obras (NomeDaObra, Regiao, estado_id, cidade_id, ClassificacaoDaObra, Status, DataDeInicio, DataDeEntrega, Orcamento, EngResponsavel, DescricaoDaObra, Imagens)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (nome_da_obra, regiao_id, estado_id, cidade_id, classificacao_id, status_id, data_inicio_obj, data_termino_obj, orcamento_utilizado, engenheiro_responsavel, descricao, json.dumps(imagem_paths)))

            db.commit()

            return jsonify({
                'success': True,
                'message': 'Obra cadastrada com sucesso!',
            }), 201
        except Exception as e:
            print("Erro ao adicionar obra:", str(e))
            traceback.print_exc()  # Imprime o traceback completo no console
            return jsonify({'error': str(e)}), 500
        finally:
            if cursor:
                cursor.close()
            if db:
                db.close()

        return jsonify({
            'success': True,
            'message': 'Obra cadastrada com sucesso!',
        }), 201
    except Exception as e:
        print("Erro ao adicionar obra:", str(e))
        traceback.print_exc()  # Imprime o traceback completo no console
        return jsonify({'error': str(e)}), 500
    
def obter_status_do_banco():
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id, Classificacao FROM statusdaobra")  # Correção aqui
        status_options = cursor.fetchall()
        cursor.close()
        conn.close()
        return status_options
    except Exception as e:
        print("Erro ao obter status do banco de dados:", e)
        return []

@app.route('/status', methods=['GET'])
def get_status():
    try:
        status_options = obter_status_do_banco()
        return jsonify(status_options), 200
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/obras-recentes', methods=['GET'])
def obras_recentes():
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        cursor.execute("""
            SELECT id, DescricaoDaObra, Imagens
            FROM obras
            ORDER BY id DESC
            LIMIT 6
        """)
        obras = cursor.fetchall()

        cursor.close()
        db.close()

        # Modificando a resposta para retornar as obras
        return jsonify({
            'obras': obras
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/obras-editar', methods=['GET'])
def obras_editar():
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        print("Executando a consulta SQL...")  # Adicionado print()
        cursor.execute("""
            SELECT Id, Imagens
            FROM obras
            ORDER BY Id DESC
        """)
        obras = cursor.fetchall()

        print("Dados recuperados do banco de dados:")  # Adicionado print()
        print(obras)  # Adicionado print()

        cursor.close()
        db.close()

        print("Dados que serão enviados na resposta da API:")  # Adicionado print()
        print(obras)  # Adicionado print()

        return jsonify({
            'obras': obras
        })
    except Exception as e:
        print("Erro:", e)  # Adicionado print()
        return jsonify({'error': str(e)}), 500


@app.route('/uploads/<filename>')
def get_image(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


#  Rota de logout (simples confirmação)
@app.route('/logout', methods=['POST'])
def logout():
    return jsonify({'success': True, 'message': 'Logout realizado com sucesso!'})

@app.route('/tela_pos_login')
def tela_pos_login():
    return send_from_directory(os.path.join(app.static_folder, 'html'), 'tela_pos_login.html')

@app.route('/obras/<int:obra_id>', methods=['GET'])
def get_obra(obra_id):
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        # Modifique a consulta para retornar todos os campos
        cursor.execute("SELECT * FROM obras WHERE Id = %s", (obra_id,))
        obra = cursor.fetchone()

        cursor.close()
        db.close()

        if obra:
            return jsonify(obra)
        else:
            return jsonify({"error": "Obra não encontrada"}), 404

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/obras/<int:obra_id>', methods=['PUT'])
def atualizar_obra(obra_id):
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()

        # Obter dados do FormData
        nome_da_obra = request.form.get('NomeDaObra')
        regiao = request.form.get('Regiao')
        classificacao_da_obra = request.form.get('ClassificacaoDaObra')
        data_de_inicio = request.form.get('DataDeInicio')
        data_de_entrega = request.form.get('DataDeEntrega')
        orcamento = request.form.get('Orcamento')
        eng_responsavel = request.form.get('EngResponsavel')
        status = request.form.get('Status')
        descricao_da_obra = request.form.get('DescricaoDaObra')

        # Obter imagens existentes do banco de dados
        cursor.execute("SELECT Imagens FROM obras WHERE Id = %s", (obra_id,))
        resultado = cursor.fetchone()
        imagens_existentes = json.loads(resultado[0]) if resultado and resultado[0] else []

        # Obter novas imagens


        novas_imagens_paths = []


        imagens = request.files.getlist('imagem')
        imagem_paths = []
        print(f"Número de arquivos enviados: {len(imagens)}")

        for imagem in imagens:
            print(f"Processando arquivo: {imagem.filename}")
            if imagem:
                filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{imagem.filename}"
                imagem_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                imagem.save(imagem_path)
                imagem_paths.append(f"/uploads/{filename}")

        # Combinar imagens existentes e novas (removendo duplicatas)
        todas_imagens = list(set(imagens_existentes + novas_imagens_paths))

        # Obter imagens a serem removidas
        imagens_para_remover = json.loads(request.form.get('imagensParaRemover')) if request.form.get('imagensParaRemover') else []

        # Filtrar imagens existentes
        todas_imagens = [img for img in todas_imagens if img not in imagens_para_remover]

        # Atualizar o banco de dados
        sql = """UPDATE obras SET NomeDaObra=%s, Regiao=%s, ClassificacaoDaObra=%s,
                    DataDeInicio=%s, DataDeEntrega=%s, Orcamento=%s, EngResponsavel=%s,
                    Status=%s, DescricaoDaObra=%s, Imagens=%s WHERE Id=%s"""

        valores = (
            nome_da_obra,
            regiao,
            classificacao_da_obra,
            data_de_inicio,
            data_de_entrega,
            orcamento,
            eng_responsavel,
            status,
            descricao_da_obra,
            json.dumps(todas_imagens),
            obra_id
        )

        cursor.execute(sql, valores)
        conn.commit()
        cursor.close()
        conn.close()

        return jsonify({'mensagem': 'Obra atualizada com sucesso!'}), 200

    except Exception as e:
        print("Erro ao atualizar obra:", e)
        return jsonify({'erro': str(e)}), 500

@app.route('/obras-consultar', methods=['GET'])
def obras_consultar():
    try:
        db = get_db_connection()
        cursor = db.cursor(dictionary=True)

        print("Executando a consulta SQL...")  # Adicionado print()
        cursor.execute("""
            SELECT Id, Imagens
            FROM obras
            ORDER BY Id DESC
        """)
        obras = cursor.fetchall()

        print("Dados recuperados do banco de dados:")  # Adicionado print()
        print(obras)  # Adicionado print()

        cursor.close()
        db.close()

        print("Dados que serão enviados na resposta da API:")  # Adicionado print()
        print(obras)  # Adicionado print()

        return jsonify({
            'obras': obras
        })
    except Exception as e:
        print("Erro:", e)  # Adicionado print()
        return jsonify({'error': str(e)}), 500

from datetime import datetime, timedelta

@app.route('/obras-excel', methods=['POST'])
def obras_excel():
    try:
        if 'arquivo' not in request.files:
            return jsonify({'success': False, 'message': 'Nenhum arquivo enviado'})

        arquivo_excel = request.files['arquivo']
        if arquivo_excel.filename == '':
            return jsonify({'success': False, 'message': 'Nenhum arquivo selecionado'})

        if arquivo_excel:
            workbook = load_workbook(arquivo_excel)
            sheet = workbook.active
            obras = []
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2):
                obra = {}
                for i, cell in enumerate(row):
                    obra[headers[i]] = cell.value
                obras.append(obra)

            db = get_db_connection()
            cursor = db.cursor()

            for obra in obras:
                nome_da_obra = obra.get('NomeDaObra')
                regiao_nome = obra.get('Regiao')
                classificacaoDaObra_nome = obra.get('ClassificacaoDaObra')
                status_nome = obra.get('Status')
                data_inicio = obra.get('DataDeInicio')
                data_termino = obra.get('DataDeEntrega')
                orcamento_utilizado = obra.get('Orcamento')
                engenheiro_responsavel = obra.get('EngResponsavel')
                descricao = obra.get('DescricaoDaObra')
                imagens = obra.get('Imagens')

                # Processar imagens
                imagens = obra.get('Imagens')
                links_imagens = []
                if imagens:
                    imagens_lista = imagens.split(',')
                    for nome_imagem in imagens_lista:
                        nome_imagem = nome_imagem.strip()
                        caminho_imagem = os.path.join(os.path.dirname(arquivo_excel.filename), nome_imagem)
                        if os.path.exists(caminho_imagem):
                            nome_arquivo_unico = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{secure_filename(nome_imagem)}"
                            caminho_destino = os.path.join(app.config['UPLOAD_FOLDER'], nome_arquivo_unico)
                            os.rename(caminho_imagem, caminho_destino)
                            link_imagem = f"/uploads/{nome_arquivo_unico}"
                            links_imagens.append(link_imagem)
                            print(f"Imagem processada: {link_imagem}") # Adicione esta linha
                        else:
                            print(f"Imagem não encontrada: {nome_imagem}") # Adicione esta linha
                imagens = str(links_imagens)
                print(f"Links das imagens: {imagens}") # Adicione esta linha

                # Buscar ID da região
                cursor.execute("SELECT id FROM regioesbrasil WHERE Nome_Regiao = %s", (regiao_nome,))
                regiao_id = cursor.fetchone()
                if regiao_id:
                    regiao_id = regiao_id[0]
                else:
                    return jsonify({'success': False, 'message': f'Região "{regiao_nome}" não encontrada'})
                cursor.fetchall()
                cursor.close()
                cursor = db.cursor()

                # Buscar ID da classificação
                cursor.execute("SELECT id FROM classificacaodasobras WHERE TipoDeObra = %s", (classificacaoDaObra_nome,))
                classificacao_id = cursor.fetchone()
                if classificacao_id:
                    classificacao_id = classificacao_id[0]
                else:
                    return jsonify({'success': False, 'message': f'Classificação "{classificacaoDaObra_nome}" não encontrada'})
                cursor.fetchall()
                cursor.close()
                cursor = db.cursor()

                # Buscar ID do status
                cursor.execute("SELECT id FROM statusdaobra WHERE Classificacao = %s", (status_nome,))
                status_id = cursor.fetchone()
                if status_id:
                    status_id = status_id[0]
                else:
                    return jsonify({'success': False, 'message': f'Status "{status_nome}" não encontrado'})
                cursor.fetchall()
                cursor.close()
                cursor = db.cursor()

                # Converter números de série para objetos datetime
                if isinstance(data_inicio, int):
                    data_inicio = datetime(1899, 12, 30) + timedelta(days=data_inicio)
                if isinstance(data_termino, int):
                    data_termino = datetime(1899, 12, 30) + timedelta(days=data_termino)

                # Usar objetos datetime diretamente
                data_inicio_obj = data_inicio.date()
                data_termino_obj = data_termino.date()

                cursor.execute("""
                    INSERT INTO obras (NomeDaObra, Regiao, ClassificacaoDaObra, Status, DataDeInicio, DataDeEntrega, Orcamento, EngResponsavel, DescricaoDaObra, Imagens)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (nome_da_obra, regiao_id, classificacao_id, status_id, data_inicio_obj, data_termino_obj, orcamento_utilizado, engenheiro_responsavel, descricao, imagens))

            db.commit()
            cursor.close()
            db.close()

            return jsonify({'success': True, 'message': 'Obras anexadas com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/download-modelo-excel')
def download_modelo():
    caminho_arquivo = os.path.join(app.root_path, 'FRONT', 'arquivos', 'PlanilhaExemplo.xlsx')
    if os.path.exists(caminho_arquivo):
        return send_file(caminho_arquivo, as_attachment=True, download_name='ModeloCadastroObras.xlsx')
    else:
        return "Arquivo não encontrado", 404

@app.route('/obras/verificar-cadastro', methods=['GET'])
def verificar_cadastro_obra():
    nome = request.args.get('nome')
    data_inicio = request.args.get('dataInicio')
    print(f"\n--- /obras/verificar-cadastro (GET) --- Nome: {nome}, Data Início: {data_inicio}")
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id FROM obras WHERE NomeDaObra = %s AND DataDeInicio = %s", (nome, data_inicio))
        obra = cursor.fetchone()
        cursor.close()
        conn.close()
        if obra:
            print("Obra encontrada para verificação.")
            return jsonify({'obraExiste': True})
        else:
            print("Obra não encontrada para verificação.")
            return jsonify({'obraExiste': False})
    except Exception as e:
        print(f"Erro ao verificar cadastro da obra: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5500) # Certifique-se de que debug=True está aqui